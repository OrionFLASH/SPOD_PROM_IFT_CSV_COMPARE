#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Сравнение пар CSV-файлов из двух каталогов (ONE и TWO).
Формирование сводных листов MERGE и списка изменений COMPARE в одном Excel-файле.
Универсальная обработка любых CSV (разделитель ";", UTF-8, первая строка — заготовки).
"""

from __future__ import annotations

import csv
import logging
import multiprocessing
import os
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from typing import TYPE_CHECKING, Any, Optional

# openpyxl нужен для записи xlsx; входит во многие дистрибутивы Python/Anaconda
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# -----------------------------------------------------------------------------
# ПАРАМЕТРЫ ПРОГРАММЫ (все настройки в одном месте)
# -----------------------------------------------------------------------------

# Корневой каталог проекта (относительно него задаются IN/OUT/log)
BASE_DIR: str = os.path.dirname(os.path.abspath(__file__))

# Каталоги с исходными CSV: ONE — первый набор, TWO — второй
# Для каждой пары сравниваемых файлов в INPUT_FILES задаются file_one (каталог ONE) и file_two (каталог TWO)
FOLDER_ONE: str = "IN/ONE"
FOLDER_TWO: str = "IN/TWO"

# Какой каталог считать «текущим» (старая версия), какой — «новым» (обновлённая)
# CURRENT_FOLDER = "ONE"  → текущий = ONE, новый = TWO
# CURRENT_FOLDER = "TWO"  → текущий = TWO, новый = ONE
CURRENT_FOLDER: str = "ONE"

# Каталог для выходного Excel и подкаталог для логов
OUT_DIR: str = "OUT"
LOG_DIR: str = "log"

# Префикс имени выходного файла (к нему добавляется таймштамп)
OUTPUT_FILE_PREFIX: str = "SPOD_PROM_IFT_COMPARE"

# Кодировка и разделитель CSV
CSV_ENCODING: str = "utf-8"
CSV_DELIMITER: str = ";"

# Цвета для заголовков в MERGE: удалённые в новом файле колонки / добавленные
COLOR_HEADER_REMOVED_IN_NEW: str = "FF0000"   # красный (колонка только в текущем)
COLOR_HEADER_ADDED_IN_NEW: str = "00FF00"     # салатовый (колонка только в новом)

# Максимальная длина листа Excel (ограничение openpyxl/Excel)
EXCEL_MAX_ROW: int = 1_048_576
EXCEL_MAX_COL: int = 16_384

# Количество процессов для параллельной обработки файлов (1 = без параллелизма)
NUM_WORKERS: int = max(1, multiprocessing.cpu_count() - 1)

# Описание каждой пары файлов для сравнения: имена в каталоге ONE и TWO, лист в Excel,
# ключевые колонки и оформление. file_one — имя файла в FOLDER_ONE, file_two — в FOLDER_TWO;
# имена могут совпадать или различаться (например, разные версии одного набора данных).
# key_columns — список имён колонок, по которым строится составной ключ (уникальность строки).
INPUT_FILES: list[dict[str, Any]] = [
    {
        "file_one": "CONTEST (PROM) 19-01 v1.csv",
        "file_two": "CONTEST (PROM) 02-02 v2.csv",
        "sheet": "CONTEST-DATA",
        "key_columns": ["CONTEST_CODE"],
        "max_col_width": 120,
        "freeze": "C2",
        "col_width_mode": "AUTO",
        "min_col_width": 12,
    },
    {
        "file_one": "GROUP (PROM) 19-01 v1.csv",
        "file_two": "GROUP (PROM) 02-02 v1.csv",
        "sheet": "GROUP",
        "key_columns": ["CONTEST_CODE", "GROUP_CODE", "GROUP_VALUE"],
        "max_col_width": 20,
        "freeze": "C2",
        "col_width_mode": "AUTO",
        "min_col_width": 8,
    },
    {
        "file_one": "INDICATOR (PROM) 19-01 v1.csv",
        "file_two": "INDICATOR (PROM) 02-02 v1.csv",
        "sheet": "INDICATOR",
        "key_columns": ["CONTEST_CODE", "INDICATOR_ADD_CALC_TYPE", "INDICATOR_CODE"],
        "max_col_width": 100,
        "freeze": "B2",
        "col_width_mode": "AUTO",
        "min_col_width": 8,
    },
    {
        "file_one": "REPORT (PROM) 19-01 v1.csv",
        "file_two": "REPORT (PROM) 02-02 v0.csv",
        "sheet": "REPORT",
        "key_columns": ["MANAGER_PERSON_NUMBER", "CONTEST_CODE", "TOURNAMENT_CODE"],
        "max_col_width": 25,
        "freeze": "D2",
        "col_width_mode": "AUTO",
        "min_col_width": 10,
    },
    {
        "file_one": "REWARD (PROM) 19-01 v2.csv",
        "file_two": "REWARD (PROM) 02-02 v1.csv",
        "sheet": "REWARD",
        "key_columns": ["REWARD_CODE"],
        "max_col_width": 200,
        "freeze": "D2",
        "col_width_mode": "AUTO",
        "min_col_width": 10,
    },
    {
        "file_one": "REWARD-LINK (PROM) 19-01 v1.csv",
        "file_two": "REWARD-LINK (PROM) 02-02 v1.csv",
        "sheet": "REWARD-LINK",
        "key_columns": ["REWARD_CODE", "CONTEST_CODE", "GROUP_CODE"],
        "max_col_width": 30,
        "freeze": "A2",
        "col_width_mode": "AUTO",
        "min_col_width": 10,
    },
    {
        "file_one": "SCHEDULE (PROM) 19-01 v2.csv",
        "file_two": "SCHEDULE (PROM) 02-02 v1.csv",
        "sheet": "TOURNAMENT-SCHEDULE",
        "key_columns": ["TOURNAMENT_CODE"],
        "max_col_width": 120,
        "freeze": "B2",
        "col_width_mode": "AUTO",
        "min_col_width": 10,
    },
    {
        "file_one": "SVD_KB_DM_GAMIFICATION_ORG_UNIT_V20 - 2025.08.28.csv",
        "file_two": "SVD_KB_DM_GAMIFICATION_ORG_UNIT_V20 - 2025.08.28.csv",
        "sheet": "ORG_UNIT_V20",
        "key_columns": ["ORG_UNIT_CODE"],
        "max_col_width": 60,
        "freeze": "A2",
        "col_width_mode": "AUTO",
        "min_col_width": 10,
    },
    {
        "file_one": "USER_ROLE (PROM) 12-12 v0.csv",
        "file_two": "USER_ROLE (PROM) 12-12 v0.csv",
        "sheet": "USER_ROLE",
        "key_columns": ["RULE_NUM", "ROLE_CODE"],
        "max_col_width": 65,
        "freeze": "D2",
        "col_width_mode": "AUTO",
        "min_col_width": 12,
    },
    {
        "file_one": "USER_ROLE_SB (PROM) 12-12 v0.csv",
        "file_two": "USER_ROLE_SB (PROM) 12-12 v0.csv",
        "sheet": "USER_ROLE SB",
        "key_columns": ["RULE_NUM", "ROLE_CODE"],
        "max_col_width": 65,
        "freeze": "D2",
        "col_width_mode": "AUTO",
        "min_col_width": 12,
    },
    {
        "file_one": "employee_PROM_final_5000_2025-07-26_00-09-03.csv",
        "file_two": "employee_PROM_final_5000_2025-07-26_00-09-03.csv",
        "sheet": "EMPLOYEE",
        "key_columns": ["PERSON_NUMBER"],
        "max_col_width": 80,
        "freeze": "F2",
        "col_width_mode": "AUTO",
        "min_col_width": 15,
    },
    {
        "file_one": "gamification-employeeRewards.csv",
        "file_two": "gamification-employeeRewards.csv",
        "sheet": "LIST-REWARDS",
        "key_columns": ["Уникальный идентификатор записи"],
        "max_col_width": 40,
        "freeze": "D2",
        "col_width_mode": "AUTO",
        "min_col_width": 12,
    },
    {
        "file_one": "gamification-statistics.csv",
        "file_two": "gamification-statistics.csv",
        "sheet": "STATISTICS",
        "key_columns": ["Табельный номер"],
        "max_col_width": 25,
        "freeze": "C2",
        "col_width_mode": "AUTO",
        "min_col_width": 10,
    },
    {
        "file_one": "gamification-tournamentList.csv",
        "file_two": "gamification-tournamentList.csv",
        "sheet": "LIST-TOURNAMENT",
        "key_columns": ["Код турнира"],
        "max_col_width": 80,
        "freeze": "C2",
        "col_width_mode": "AUTO",
        "min_col_width": 12,
    },
]


# -----------------------------------------------------------------------------
# ЛОГИРОВАНИЕ
# -----------------------------------------------------------------------------

def _setup_logging(log_dir: str) -> logging.Logger:
    """
    Настройка логирования в файл и консоль.
    Уровни: INFO — основные события, DEBUG — диагностика.
    Имя файла: DEBUG_compare_ГГГГММДД_ЧЧ.log
    Формат строки DEBUG: дата время - [уровень] - сообщение [def: имя_функции]
    """
    os.makedirs(log_dir, exist_ok=True)
    log_name = datetime.now().strftime("DEBUG_compare_%Y%m%d_%H")
    log_path = os.path.join(log_dir, f"{log_name}.log")

    logger = logging.getLogger("csv_compare")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fmt_debug = "%(asctime)s - [%(levelname)s] - %(message)s [def: %(funcName)s]"
    fmt_info = "%(asctime)s - [%(levelname)s] - %(message)s"
    date_fmt = "%Y-%m-%d %H:%M:%S"

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(fmt_debug, datefmt=date_fmt))
    logger.addHandler(fh)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter(fmt_info, datefmt=date_fmt))
    logger.addHandler(ch)

    logger.debug("Логгер создан, файл: %s", log_path)
    return logger


# -----------------------------------------------------------------------------
# ЗАГРУЗКА CSV
# -----------------------------------------------------------------------------

def _resolve_file_path(base_dir: str, folder: str, file_name: str) -> Optional[str]:
    """
    Ищет файл в каталоге: сначала точное имя, затем варианты с разным регистром расширения.
    Возвращает полный путь или None.
    """
    _log = logging.getLogger("csv_compare")
    _log.debug("Этап: поиск файла folder=%s file_name=%s", folder, file_name)
    folder_path = os.path.join(base_dir, folder)
    exact = os.path.join(folder_path, file_name)
    if os.path.isfile(exact):
        _log.debug("Файл найден (точное совпадение): %s", exact)
        return exact
    name_base, ext = os.path.splitext(file_name)
    for candidate_ext in (ext, ".csv", ".CSV"):
        if candidate_ext == ext:
            continue
        candidate = os.path.join(folder_path, name_base + candidate_ext)
        if os.path.isfile(candidate):
            _log.debug("Файл найден (другой регистр расширения): %s", candidate)
            return candidate
    _log.debug("Файл не найден: %s", os.path.join(folder_path, file_name))
    return None


def load_csv(
    file_path: str,
    encoding: str = CSV_ENCODING,
    delimiter: str = CSV_DELIMITER,
    logger: Optional[logging.Logger] = None,
) -> tuple[list[str], list[dict[str, str]]]:
    """
    Загружает CSV: первая строка — заголовки, остальные — данные.
    Возвращает (список_заголовков, список_строк_как_словарь).
    """
    if logger:
        logger.debug("Этап: чтение CSV path=%s", file_path)
    with open(file_path, "r", encoding=encoding, newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)
    if not rows:
        if logger:
            logger.debug("CSV пустой, возврат ([], [])")
        return [], []
    headers = [str(h).strip() for h in rows[0]]
    data: list[dict[str, str]] = []
    for r in rows[1:]:
        row_dict: dict[str, str] = {}
        for i, h in enumerate(headers):
            row_dict[h] = r[i].strip() if i < len(r) else ""
        data.append(row_dict)
    if logger:
        logger.debug("CSV загружен: заголовков=%d строк_данных=%d", len(headers), len(data))
    return headers, data


def make_key(row: dict[str, str], key_columns: list[str]) -> tuple[str, ...]:
    """Строит кортеж значений ключевых колонок для строки (составной ключ)."""
    return tuple(str(row.get(k, "")) for k in key_columns)


def build_index(
    headers: list[str],
    data: list[dict[str, str]],
    key_columns: list[str],
    logger: Optional[logging.Logger] = None,
) -> dict[tuple[str, ...], dict[str, str]]:
    """Строит словарь: ключ -> строка (для быстрого поиска по ключу)."""
    result: dict[tuple[str, ...], dict[str, str]] = {}
    for row in data:
        k = make_key(row, key_columns)
        result[k] = dict(row)
    if logger:
        logger.debug("Этап: build_index строк=%d уникальных_ключей=%d", len(data), len(result))
    return result


def find_value_diff(old_val: str, new_val: str) -> str:
    """
    Пытается указать место изменения внутри значения (для COMPARE).
    Возвращает краткое описание: например, «символы 5–10» или «целиком».
    """
    if old_val == new_val:
        return ""
    if not old_val or not new_val:
        return "целиком"
    o, n = old_val, new_val
    start = 0
    while start < min(len(o), len(n)) and o[start] == n[start]:
        start += 1
    end_o = len(o)
    end_n = len(n)
    while end_o > start and end_n > start and o[end_o - 1] == n[end_n - 1]:
        end_o -= 1
        end_n -= 1
    if start == 0 and end_o == len(o) and end_n == len(n):
        return "целиком"
    return f"приблизительно символы {start + 1}–{end_o} (старое) / {start + 1}–{end_n} (новое)"


# -----------------------------------------------------------------------------
# СЛИЯНИЕ И СРАВНЕНИЕ ОДНОЙ ПАРЫ ФАЙЛОВ
# -----------------------------------------------------------------------------

@dataclass
class MergeCompareResult:
    """Результат обработки одной пары файлов: данные для листов MERGE и COMPARE."""

    sheet_name: str
    # Заголовки: общий порядок колонок для MERGE (сначала текущий, потом добавленные из нового)
    headers_merge: list[str]
    # Для каждой колонки в headers_merge: True = только в текущем (красный), False = в обоих, None = только в новом (салатовый)
    header_flags: list[Optional[bool]]
    rows_merge: list[dict[str, str]]
    # COMPARE: список записей {key, column, value_current, value_new, location_in_value} (исходный)
    compare_rows: list[dict[str, Any]]
    # COMPARE: группы по ключу для вывода в 3 строки: [{change_no, key, row_one, row_two, changes}, ...]
    compare_groups: list[dict[str, Any]]
    # Параметры оформления листа
    max_col_width: int
    freeze: str
    min_col_width: int
    # Данные для листов ONE и TWO (заголовки + строки)
    headers_one: list[str]
    rows_one: list[dict[str, str]]
    headers_two: list[str]
    rows_two: list[dict[str, str]]
    # Ключевые колонки (порядок для вывода ключа в COMPARE)
    key_columns: list[str]


def process_one_file(
    config: dict[str, Any],
    path_current: Optional[str],
    path_new: Optional[str],
    current_folder_name: str,
    new_folder_name: str,
    key_columns: list[str],
    logger: Optional[logging.Logger] = None,
) -> MergeCompareResult:
    """
    Загружает два CSV, строит множество уникальных ключей, заполняет MERGE
    (текущий + перезапись из нового где есть изменения) и список COMPARE.
    """
    sheet_name = config["sheet"]
    if logger:
        logger.debug("Этап: process_one_file sheet=%s path_current=%s path_new=%s", sheet_name, path_current, path_new)
    result = MergeCompareResult(
        sheet_name=sheet_name,
        headers_merge=[],
        header_flags=[],
        rows_merge=[],
        compare_rows=[],
        compare_groups=[],
        max_col_width=config.get("max_col_width", 50),
        freeze=config.get("freeze", "A2"),
        min_col_width=config.get("min_col_width", 10),
        headers_one=[],
        rows_one=[],
        headers_two=[],
        rows_two=[],
        key_columns=key_columns,
    )

    # Загрузка текущего и нового файлов
    if path_current:
        result.headers_one, result.rows_one = load_csv(path_current, logger=logger)
    else:
        result.headers_one, result.rows_one = [], []

    if path_new:
        result.headers_two, result.rows_two = load_csv(path_new, logger=logger)
    else:
        result.headers_two, result.rows_two = [], []

    # Проверка наличия ключевых колонок в текущем файле
    for k in key_columns:
        if result.headers_one and k not in result.headers_one:
            if logger:
                logger.warning("Ключевая колонка '%s' отсутствует в текущем файле %s", k, path_current or "")

    # Порядок колонок в MERGE: сначала все из текущего файла, затем колонки,
    # которые есть только в новом (добавленные). Флаги окраски заголовка:
    # False — колонка в обоих файлах; True — только в текущем (красный);
    # None — только в новом (салатовый).
    set_current = set(result.headers_one)
    set_new = set(result.headers_two)
    headers_merge: list[str] = list(result.headers_one)
    header_flags: list[Optional[bool]] = [False] * len(headers_merge)
    for col in result.headers_two:
        if col not in set_current:
            headers_merge.append(col)
            header_flags.append(None)  # только в новом — салатовый
    for i, col in enumerate(result.headers_one):
        if col not in set_new:
            header_flags[i] = True  # только в текущем — красный

    idx_current = build_index(result.headers_one, result.rows_one, key_columns, logger=logger)
    idx_new = build_index(result.headers_two, result.rows_two, key_columns, logger=logger)
    all_keys = sorted(set(idx_current.keys()) | set(idx_new.keys()))
    if logger:
        logger.debug("Этап: merge ключей всего=%d в_текущем=%d в_новом=%d колонок_merge=%d", len(all_keys), len(idx_current), len(idx_new), len(headers_merge))

    # Сборка MERGE: для каждого уникального ключа берём строку из текущего,
    # затем перезаписываем значениями из нового там, где они есть; отличия
    # фиксируем в compare_rows для листа COMPARE. Ключи с правками — для колонки CHANGE.
    keys_with_changes: set[tuple[str, ...]] = set()
    for key in all_keys:
        row_current = idx_current.get(key)
        row_new = idx_new.get(key)
        merged_row: dict[str, str] = {}
        for col in headers_merge:
            val_cur = (row_current or {}).get(col, "")
            val_new = (row_new or {}).get(col, "")
            merged_row[col] = val_cur
            if row_new is not None and col in row_new:
                if val_cur != val_new:
                    merged_row[col] = val_new
                    keys_with_changes.add(key)
                    result.compare_rows.append({
                        "key": key,
                        "column": col,
                        "value_current": val_cur,
                        "value_new": val_new,
                        "location_in_value": find_value_diff(val_cur, val_new),
                    })
            elif row_current is not None:
                merged_row[col] = val_cur
        # Колонка CHANGE в конце листа: NEW — ключа не было в ONE, DEL — ключа нет в TWO, CHANGE — правки по строке
        if row_current is None:
            merged_row["CHANGE"] = "NEW"
        elif row_new is None:
            merged_row["CHANGE"] = "DEL"
        elif key in keys_with_changes:
            merged_row["CHANGE"] = "CHANGE"
        else:
            merged_row["CHANGE"] = "-"
        result.rows_merge.append(merged_row)
    result.headers_merge = headers_merge + ["CHANGE"]
    result.header_flags = header_flags + [False]

    # Группировка изменений по ключу для листа COMPARE: три строки на правку (ONE, TWO, CHANGE).
    # Учитываем: (1) ключи с изменением значений колонок — CHANGE; (2) ключи только в ONE — DEL; (3) ключи только в TWO — NEW.
    by_key: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for rec in result.compare_rows:
        by_key[rec["key"]].append(rec)
    keys_change = set(by_key.keys())
    keys_only_one = set(idx_current.keys()) - set(idx_new.keys())
    keys_only_two = set(idx_new.keys()) - set(idx_current.keys())
    all_compare_keys = sorted(keys_change | keys_only_one | keys_only_two)
    result.compare_groups = []
    for change_no, key in enumerate(all_compare_keys, start=1):
        row_one = dict(idx_current.get(key) or {})
        row_two = dict(idx_new.get(key) or {})
        changes: dict[str, str] = {}
        if key in keys_only_one:
            change_type = "DEL"
        elif key in keys_only_two:
            change_type = "NEW"
        else:
            change_type = "CHANGE"
            recs = by_key[key]
            for r in recs:
                col = r["column"]
                v_cur = r.get("value_current", "")
                v_new = r.get("value_new", "")
                loc = r.get("location_in_value", "")
                if len(str(v_cur)) + len(str(v_new)) <= 100:
                    part = f"было: {v_cur} → стало: {v_new}"
                else:
                    part = "было → стало"
                if loc:
                    changes[col] = f"{part} ({loc})"
                else:
                    changes[col] = part
        result.compare_groups.append({
            "change_no": change_no,
            "key": key,
            "row_one": row_one,
            "row_two": row_two,
            "changes": changes,
            "change_type": change_type,
        })

    if logger:
        logger.debug("process_one_file завершён sheet=%s строк_merge=%d групп_compare=%d", sheet_name, len(result.rows_merge), len(result.compare_groups))
    return result


def _worker_process_file(
    args: tuple[dict[str, Any], str, str, str, str, list[str]],
) -> MergeCompareResult:
    """
    Обёртка для вызова process_one_file из пула процессов (без передачи logger).
    """
    config, path_current, path_new, cur_name, new_name, key_cols = args
    return process_one_file(
        config, path_current, path_new, cur_name, new_name, key_cols, logger=None
    )


# -----------------------------------------------------------------------------
# ЗАПИСЬ EXCEL
# -----------------------------------------------------------------------------

def _write_sheet_data(
    ws: Worksheet,
    headers: list[str],
    rows: list[dict[str, str]],
    bold_header: bool = True,
    header_fill: Optional[list[Optional[bool]]] = None,
    max_col_width: int = 50,
    min_col_width: int = 8,
    freeze: str = "A2",
    auto_filter: bool = True,
    logger: Optional[logging.Logger] = None,
) -> None:
    """
    Записывает на лист заголовки и данные, форматирует первую строку (жирный),
    при необходимости раскрашивает заголовки (header_fill: True=красный, None=салатовый),
    устанавливает закрепление и автофильтр.
    """
    if logger:
        logger.debug("Этап: запись листа title=%s колонок=%d строк=%d", getattr(ws, "title", "?"), len(headers), len(rows))
    font_bold = Font(bold=True)
    fill_red = PatternFill(start_color=COLOR_HEADER_REMOVED_IN_NEW, end_color=COLOR_HEADER_REMOVED_IN_NEW, fill_type="solid")
    fill_green = PatternFill(start_color=COLOR_HEADER_ADDED_IN_NEW, end_color=COLOR_HEADER_ADDED_IN_NEW, fill_type="solid")

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font_bold
        if header_fill is not None and col_idx <= len(header_fill):
            flag = header_fill[col_idx - 1]
            if flag is True:
                cell.fill = fill_red
            elif flag is None:
                cell.fill = fill_green

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, h in enumerate(headers, start=1):
            val = row.get(h, "")
            if isinstance(val, str) and len(val) > 32767:
                val = val[:32767]
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Ширина колонок
    for col_idx in range(1, len(headers) + 1):
        try:
            max_len = min(max_col_width, max(
                min_col_width,
                *(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, min(ws.max_row + 1, 1024)))
            ))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(min(max_len + 1, max_col_width), min_col_width)
        except Exception:
            ws.column_dimensions[get_column_letter(col_idx)].width = min_col_width

    if freeze:
        ws.freeze_panes = freeze
    if auto_filter and rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def _write_compare_sheet(
    ws: Worksheet,
    compare_groups: list[dict[str, Any]],
    headers_merge: list[str],
    key_columns: list[str],
    bold_header: bool = True,
    logger: Optional[logging.Logger] = None,
) -> None:
    """
    Пишет лист COMPARE: для каждой правки три строки —
    1) ONE: ключ и все значения из текущего файла;
    2) TWO: тот же ключ и все значения из нового файла;
    3) CHANGE: ключ (для привязки к строке) + пометки по колонкам (что поменялось) или DEL/NEW.
    Для DEL (строка только в ONE): ONE — полная строка, TWO — пусто, CHANGE — ключ + «DEL».
    Для NEW (строка только в TWO): ONE — пусто, TWO — полная строка, CHANGE — ключ + «NEW».
    Колонки: №, откуда (ONE/TWO/CHANGE), затем все колонки из MERGE.
    """
    if logger:
        logger.debug("Этап: запись листа COMPARE title=%s групп_правок=%d", getattr(ws, "title", "?"), len(compare_groups))
    headers_c = ["№", "откуда"] + list(headers_merge)
    font_bold = Font(bold=True)
    for col_idx, h in enumerate(headers_c, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font_bold
    row_idx = 2
    for group in compare_groups:
        change_no = group["change_no"]
        row_one = group.get("row_one") or {}
        row_two = group.get("row_two") or {}
        changes = group.get("changes") or {}
        change_type = group.get("change_type", "CHANGE")
        key = group.get("key", ())
        # Строка 1: ONE — полная строка из текущего файла
        ws.cell(row=row_idx, column=1, value=change_no)
        ws.cell(row=row_idx, column=2, value="ONE")
        for col_idx, col in enumerate(headers_merge, start=3):
            val = row_one.get(col, "")
            if isinstance(val, str) and len(val) > 32767:
                val = val[:32767]
            ws.cell(row=row_idx, column=col_idx, value=val)
        row_idx += 1
        # Строка 2: TWO — полная строка из нового файла
        ws.cell(row=row_idx, column=1, value=change_no)
        ws.cell(row=row_idx, column=2, value="TWO")
        for col_idx, col in enumerate(headers_merge, start=3):
            val = row_two.get(col, "")
            if isinstance(val, str) and len(val) > 32767:
                val = val[:32767]
            ws.cell(row=row_idx, column=col_idx, value=val)
        row_idx += 1
        # Строка 3: CHANGE — ключ (чтобы было понятно, к какой строке относится) + пометки по колонкам или DEL/NEW
        first_non_key_col = next((c for c in headers_merge if c not in key_columns), None)
        col_for_del_new = first_non_key_col if first_non_key_col is not None else (headers_merge[0] if headers_merge else None)
        ws.cell(row=row_idx, column=1, value=change_no)
        ws.cell(row=row_idx, column=2, value="CHANGE")
        for col_idx, col in enumerate(headers_merge, start=3):
            if col in key_columns:
                idx = key_columns.index(col)
                val = key[idx] if idx < len(key) else ""
            elif change_type == "DEL" and col == col_for_del_new:
                val = "DEL"
            elif change_type == "NEW" and col == col_for_del_new:
                val = "NEW"
            else:
                val = changes.get(col, "")
            ws.cell(row=row_idx, column=col_idx, value=val)
        row_idx += 1
    num_cols = len(headers_c)
    for col_idx in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(30, 100)
    ws.freeze_panes = "B2"
    if compare_groups:
        last_row = 1 + 3 * len(compare_groups)
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{last_row}"


def _write_excel_in(
    results: list[MergeCompareResult],
    output_path: str,
    current_folder_name: str,
    new_folder_name: str,
    logger: Optional[logging.Logger] = None,
) -> None:
    """Создаёт xlsx с исходными данными: для каждого результата — листы ONE и TWO."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("Модуль openpyxl не найден. Установите: conda install openpyxl")
    if logger:
        logger.debug("Этап: запись файла IN (ONE/TWO) путь=%s", output_path)
    print("  Запись IN (исходные ONE/TWO)...", flush=True)
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    first = True
    for res in results:
        title_one = f"{res.sheet_name} {current_folder_name}"
        ws_one = wb.create_sheet(title=title_one) if not first else default_sheet
        if first:
            ws_one.title = title_one
            first = False
        _write_sheet_data(
            ws_one, res.headers_one, res.rows_one,
            bold_header=True, header_fill=None,
            max_col_width=res.max_col_width, min_col_width=res.min_col_width,
            freeze=res.freeze, auto_filter=True, logger=logger,
        )
        title_two = f"{res.sheet_name} {new_folder_name}"
        ws_two = wb.create_sheet(title=title_two)
        _write_sheet_data(
            ws_two, res.headers_two, res.rows_two,
            bold_header=True, header_fill=None,
            max_col_width=res.max_col_width, min_col_width=res.min_col_width,
            freeze=res.freeze, auto_filter=True, logger=logger,
        )
    wb.save(output_path)
    if logger:
        logger.info("Файл сохранён (исходные данные): %s", output_path)


def _write_excel_merge(
    results: list[MergeCompareResult],
    output_path: str,
    logger: Optional[logging.Logger] = None,
) -> None:
    """Создаёт xlsx только с листами MERGE."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("Модуль openpyxl не найден. Установите: conda install openpyxl")
    if logger:
        logger.debug("Этап: запись файла MERGE путь=%s", output_path)
    print("  Запись MERGE (слияние по ключу)...", flush=True)
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    first = True
    for res in results:
        title_merge = f"MERGE {res.sheet_name}"
        ws = wb.create_sheet(title=title_merge) if not first else default_sheet
        if first:
            ws.title = title_merge
            first = False
        _write_sheet_data(
            ws, res.headers_merge, res.rows_merge,
            bold_header=True, header_fill=res.header_flags,
            max_col_width=res.max_col_width, min_col_width=res.min_col_width,
            freeze=res.freeze, auto_filter=True, logger=logger,
        )
    wb.save(output_path)
    if logger:
        logger.info("Файл сохранён (MERGE): %s", output_path)


def _write_excel_compare(
    results: list[MergeCompareResult],
    output_path: str,
    logger: Optional[logging.Logger] = None,
) -> None:
    """Создаёт xlsx только с листами COMPARE."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("Модуль openpyxl не найден. Установите: conda install openpyxl")
    if logger:
        logger.debug("Этап: запись файла COMPARE путь=%s", output_path)
    print("  Запись COMPARE (список изменений)...", flush=True)
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    first = True
    for res in results:
        title_compare = f"COMPARE {res.sheet_name}"
        ws = wb.create_sheet(title=title_compare) if not first else default_sheet
        if first:
            ws.title = title_compare
            first = False
        _write_compare_sheet(ws, res.compare_groups, res.headers_merge, res.key_columns, bold_header=True, logger=logger)
    wb.save(output_path)
    if logger:
        logger.info("Файл сохранён (COMPARE): %s", output_path)


def write_excel_three_files(
    results: list[MergeCompareResult],
    out_path_dir: str,
    output_file_prefix: str,
    timestamp: str,
    current_folder_name: str,
    new_folder_name: str,
    logger: Optional[logging.Logger] = None,
) -> tuple[str, str, str]:
    """
    Создаёт три xlsx-файла: IN (исходные ONE/TWO), MERGE, COMPARE.
    Имена: {тип}_{префикс}_{таймштамп}.xlsx.
    Возвращает (путь_in, путь_merge, путь_compare).
    """
    if logger:
        logger.debug("Этап: write_excel_three_files начало результатов=%d", len(results))
    path_in = os.path.join(out_path_dir, f"IN_{output_file_prefix}_{timestamp}.xlsx")
    path_merge = os.path.join(out_path_dir, f"MERGE_{output_file_prefix}_{timestamp}.xlsx")
    path_compare = os.path.join(out_path_dir, f"COMPARE_{output_file_prefix}_{timestamp}.xlsx")
    _write_excel_in(results, path_in, current_folder_name, new_folder_name, logger=logger)
    _write_excel_merge(results, path_merge, logger=logger)
    _write_excel_compare(results, path_compare, logger=logger)
    if logger:
        logger.debug("write_excel_three_files завершён: сохранены IN, MERGE, COMPARE")
    return path_in, path_merge, path_compare


# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------

def main() -> int:
    """Точка входа: настройка логов, пути, запуск параллельной обработки и запись Excel."""
    base = BASE_DIR
    log_dir = os.path.join(base, LOG_DIR)
    logger = _setup_logging(log_dir)
    logger.info("Старт сравнения CSV: каталоги ONE и TWO")

    if not OPENPYXL_AVAILABLE:
        logger.error("Требуется openpyxl. В среде Anaconda: conda install openpyxl")
        return 1

    folder_one_path = os.path.join(base, FOLDER_ONE)
    folder_two_path = os.path.join(base, FOLDER_TWO)
    out_path_dir = os.path.join(base, OUT_DIR)
    os.makedirs(out_path_dir, exist_ok=True)

    if CURRENT_FOLDER == "ONE":
        current_folder_name, new_folder_name = "ONE", "TWO"
    else:
        current_folder_name, new_folder_name = "TWO", "ONE"

    # Проверка наличия всех указанных файлов в каталогах ONE и TWO
    logger.debug("Этап: проверка наличия всех файлов в каталогах ONE и TWO")
    missing: list[tuple[str, str]] = []  # (каталог, имя_файла)
    for cfg in INPUT_FILES:
        path_one = _resolve_file_path(base, FOLDER_ONE, cfg["file_one"])
        path_two = _resolve_file_path(base, FOLDER_TWO, cfg["file_two"])
        if path_one is None:
            missing.append(("ONE", cfg["file_one"]))
        if path_two is None:
            missing.append(("TWO", cfg["file_two"]))

    if missing:
        msg_lines = ["Обнаружены отсутствующие файлы. Работа остановлена."]
        for folder, file_name in missing:
            msg_lines.append(f"  — Каталог {folder}: {file_name}")
        full_msg = "\n".join(msg_lines)
        logger.error(full_msg)
        logger.debug("Список отсутствующих файлов: %s", missing)
        print(full_msg, file=sys.stderr)
        return 1

    logger.info("Проверка файлов: все указанные файлы найдены")
    print("  Все файлы на месте. Подготовка списка пар...", flush=True)
    logger.debug("Этап: подготовка списка файлов (пути ONE/TWO, текущий=%s)", CURRENT_FOLDER)
    tasks: list[tuple[dict[str, Any], Optional[str], Optional[str], str, str, list[str]]] = []
    for cfg in INPUT_FILES:
        file_one = cfg["file_one"]
        file_two = cfg["file_two"]
        key_columns = cfg.get("key_columns") or []
        # Путь к «текущему» файлу: в каталоге ONE или TWO в зависимости от CURRENT_FOLDER
        path_cur = _resolve_file_path(
            base,
            FOLDER_ONE if CURRENT_FOLDER == "ONE" else FOLDER_TWO,
            file_one if CURRENT_FOLDER == "ONE" else file_two,
        )
        # Путь к «новому» файлу: в противоположном каталоге
        path_new = _resolve_file_path(
            base,
            FOLDER_TWO if CURRENT_FOLDER == "ONE" else FOLDER_ONE,
            file_two if CURRENT_FOLDER == "ONE" else file_one,
        )
        tasks.append((cfg, path_cur, path_new, current_folder_name, new_folder_name, key_columns))

    logger.debug("Подготовка завершена: пар_файлов=%d режим=%s", len(tasks), "последовательно" if NUM_WORKERS <= 1 else f"параллельно (процессов={NUM_WORKERS})")
    n_tasks = len(tasks)
    logger.info("Обработка файлов (%d пар)", n_tasks)
    print(f"Обработка файлов: {n_tasks} пар.", flush=True)

    # Обработка: параллельно или последовательно
    if NUM_WORKERS <= 1:
        results = []
        for i, (cfg, path_cur, path_new, cur_name, new_name, key_cols) in enumerate(tasks, start=1):
            sheet_name = cfg["sheet"]
            file_one = cfg.get("file_one", "")
            file_two = cfg.get("file_two", "")
            print(f"  [{i}/{n_tasks}] Обработка: {sheet_name} ({file_one} / {file_two})", flush=True)
            res = process_one_file(cfg, path_cur, path_new, cur_name, new_name, key_cols, logger=logger)
            results.append(res)
            print(f"         строк MERGE: {len(res.rows_merge)}, изменений: {len(res.compare_groups)}", flush=True)
        print("Обработка всех пар завершена.", flush=True)
    else:
        print(f"  Параллельный запуск: {NUM_WORKERS} процессов, {n_tasks} пар.", flush=True)
        worker_args = [
            (cfg, path_cur, path_new, cur_name, new_name, key_cols)
            for cfg, path_cur, path_new, cur_name, new_name, key_cols in tasks
        ]
        with multiprocessing.Pool(processes=NUM_WORKERS) as pool:
            results = pool.map(_worker_process_file, worker_args)
        print("Параллельная обработка завершена.", flush=True)

    logger.debug("Обработка завершена: результатов=%d", len(results))
    logger.info("Запись результата в Excel (3 файла: IN, MERGE, COMPARE)")
    print("Запись результата в Excel (3 файла: IN, MERGE, COMPARE)...", flush=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H-%M")
    path_in, path_merge, path_compare = write_excel_three_files(
        results,
        out_path_dir,
        OUTPUT_FILE_PREFIX,
        timestamp,
        current_folder_name,
        new_folder_name,
        logger=logger,
    )
    logger.info("Готово. Выходные файлы: IN=%s MERGE=%s COMPARE=%s", path_in, path_merge, path_compare)
    print("Готово. Файлы сохранены в каталог OUT.", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
